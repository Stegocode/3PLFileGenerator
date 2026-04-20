# -*- coding: utf-8 -*-
"""
HUB Flat File Generator  v5
============================
Transforms a batch invoice from a retail appliance dealer into a HUB Group
delivery flat file, with automatic mileage pricing, Monday.com crate-status
lookup, multifamily floor-charge prompts, and a per-stop financial report.

Outputs
-------
  <PREFIX>_MM.DD.YY.xlsx          — OrderData sheet, ready to upload to HUB
  <PREFIX>_Financial_MMDDYYYY.xlsx — 3-sheet P&L report (By Stop / Service / Product)

Supported environments
----------------------
  Google Colab  — reads credentials from Colab Secrets; uploads files interactively
  Local Python  — reads credentials from .env; auto-detects files in CONTENT_DIR

HOW TO USE (Google Colab)
  1. Run Cell 1 (Chrome / drive mount setup)
  2. Add ORS_API_KEY and MONDAY_API_TOKEN to Colab Secrets
  3. Run Cell 2 (this script)
  4. Upload: bulk-invoice .xlsx, model-inventory .csv, orders-detail .csv
     Optional: serial-number-inventory .csv for unit-level cost accuracy
  5. Answer 1–2 prompts per multifamily order
  6. Script downloads finished flat file + financial report

External APIs
-------------
  Open Route Service  https://openrouteservice.org  (free, 2 000 req/day)
  Monday.com GraphQL  https://api.monday.com/v2
"""

# ── Standard library ────────────────────────────────────────────
import json
import math
import os
import shutil
import time
from datetime import datetime
from pathlib import Path

# ── Third-party ─────────────────────────────────────────────────
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Colab compatibility shim ─────────────────────────────────────
try:
    from google.colab import files as _colab_files
    _IS_COLAB = True
except ImportError:
    _IS_COLAB = False
    class _colab_files:                          # noqa: N801
        @staticmethod
        def download(path):
            print(f"  Saved: {path}")


# ══════════════════════════════════════════════════════════════════
# CONFIGURATION — edit these values to adapt to a different client
# ══════════════════════════════════════════════════════════════════
CONFIG = {
    # HUB carrier location code assigned to this account
    "location_code": "BSC1",

    # Warehouse address used as the origin for all mileage calculations
    "warehouse_address": "4600 NW St Helens Rd, Portland, OR 97210",

    # Prefix used in output file names  (e.g. "Acme" → Acme_04.18.26.xlsx)
    "output_prefix": "HUB",

    # Monday.com Delivery Scheduler board ID
    "delivery_scheduler_board_id": 8639744112,

    # Monday.com column IDs for the Delivery Scheduler board
    "monday_crate_col_id": "color_mkns94j8",   # "HOW ARE WE DELIVERING THIS?"
    "monday_date_col_id":  "date7__1",          # "REQUESTED DELIVERY DATE"

    # HUB order template path (relative to this script, or absolute)
    "template_filename": "HubGroup-OrderTemplate.xltx",

    # Colab-specific: path inside /content where the template lives
    "template_colab_dir": "/content/FlatFileGenerator",

    # Local-specific: working directory when not running in Colab
    # Override with the CONTENT_DIR environment variable if needed
    "local_content_dir": r"C:\Users\scottt\Documents\FlatFileGenerator",

    # Local-specific: export directory for finished files
    # Override with the EXPORT_DIR environment variable if needed
    "local_export_dir": r"C:\Users\scottt\Documents\FlatFileGenerator\Output",
}

# ── Mileage / fuel rate constants ──────────────────────────────
MIN_CHARGE      = 950.00    # per-truck daily minimum billed to HUB
FLOOR_CHARGE    = 18.70     # per piece per floor above floor 3
MILEAGE_CHARGE  = 30.00     # extra charge per stop, 30–125 mi
MILEAGE_FREE_MAX = 30       # miles below which no mileage charge applies
MILEAGE_MAX     = 125       # miles above which charge is flagged for review
FUEL_SHORT      = 6.60      # fuel surcharge per stop, under 30 mi
FUEL_LONG       = 9.90      # fuel surcharge per stop, 30+ mi

# ── Crate / uncrate rates (Monday → X021 / X022) ──────────────
CRATE_CHARGE_IN_BOX     = 10.20   # X021 per piece
CRATE_CHARGE_OUT_OF_BOX = 20.40   # X022 per piece

# ── Part categories in the inventory CSV ──────────────────────
PART_CATEGORIES = {"Appliance Accessories and Parts", "Laundry Pedestals"}


# ══════════════════════════════════════════════════════════════════
# SERVICE CODE TABLE
# Rates sourced from carrier rate sheet.  Add new codes here as
# they are negotiated; set charge/margin to 0.00 until confirmed.
# ══════════════════════════════════════════════════════════════════
SERVICE_CODES = {
    'X001':  {'material_type': 'SERVICE', 'description': 'White Glove',                                        'product_category': 'Service', 'charge': 80.00,  'margin': 45.00},
    'X015':  {'material_type': 'SERVICE', 'description': 'Attempted Delivery/Pickup',                          'product_category': 'Service', 'charge': 88.00,  'margin': 37.00},
    'X018':  {'material_type': 'SERVICE', 'description': 'Additional Drop Point / Threshold',                  'product_category': 'Service', 'charge': 33.00,  'margin':  0.00},
    'X021':  {'material_type': 'SERVICE', 'description': 'Crated Piece Delivered [Builder]',                   'product_category': 'Service', 'charge': 10.20,  'margin':  0.00},
    'X022':  {'material_type': 'SERVICE', 'description': 'Uncrated Piece Delivered [Builder]',                 'product_category': 'Service', 'charge': 20.40,  'margin':  0.00},
    'X029':  {'material_type': 'SERVICE', 'description': 'Appliance Anti-Tip',                                 'product_category': 'Service', 'charge':  5.50,  'margin':  0.00},
    'X100':  {'material_type': 'SERVICE', 'description': 'Haul Away/Recycle Appliance',                        'product_category': 'Service', 'charge': 22.00,  'margin':  3.00},
    'X103A': {'material_type': 'SERVICE', 'description': 'Reverse Door Swing of Appliance (A)',                'product_category': 'Service', 'charge': 27.50,  'margin':  0.00},
    'X103B': {'material_type': 'SERVICE', 'description': 'Reverse Door Swing of Appliance (B)',                'product_category': 'Service', 'charge': 27.50,  'margin':  0.00},
    'X103':  {'material_type': 'SERVICE', 'description': 'Reverse Door Swing of Appliance',                    'product_category': 'Service', 'charge': 27.50,  'margin':  0.00},
    'X151':  {'material_type': 'SERVICE', 'description': 'Install Refrigerator',                               'product_category': 'Service', 'charge': 18.75,  'margin': -18.75},
    'X152':  {'material_type': 'SERVICE', 'description': 'Install Ice Maker',                                  'product_category': 'Service', 'charge': 16.50,  'margin': 13.50},
    'X154':  {'material_type': 'SERVICE', 'description': 'Install Built-In Refrigerator',                      'product_category': 'Service', 'charge': 165.00, 'margin': 85.00},
    'X156':  {'material_type': 'SERVICE', 'description': 'Install Professional Refrigerator',                  'product_category': 'Service', 'charge': 350.00, 'margin': 75.00},
    'X157':  {'material_type': 'SERVICE', 'description': 'PRO Trim Kit for FS Refrigerator',                   'product_category': 'Service', 'charge': 150.00, 'margin':  5.00},
    'X159':  {'material_type': 'SERVICE', 'description': 'Install Ice Machine - Large/Commercial',              'product_category': 'Service', 'charge':  65.50, 'margin': 59.50},
    'X164':  {'material_type': 'SERVICE', 'description': 'Install Undercounter Refrigerator',                  'product_category': 'Service', 'charge': 125.00, 'margin': 38.00},
    'X166':  {'material_type': 'SERVICE', 'description': 'Install Beverage Center',                            'product_category': 'Service', 'charge':  12.00, 'margin': 38.00},
    'X201':  {'material_type': 'SERVICE', 'description': 'Install Electric Range',                             'product_category': 'Service', 'charge':  24.75, 'margin': 100.25},
    'X204':  {'material_type': 'SERVICE', 'description': 'Install Gas Range',                                  'product_category': 'Service', 'charge':  32.75, 'margin': 92.25},
    'X208':  {'material_type': 'SERVICE', 'description': 'Install Electric Double Wall Oven',                  'product_category': 'Service', 'charge': 162.50, 'margin': 87.50},
    'X212':  {'material_type': 'SERVICE', 'description': 'Install Gas Double Wall Oven',                       'product_category': 'Service', 'charge': 195.50, 'margin': 54.50},
    'X214':  {'material_type': 'SERVICE', 'description': 'Install Electric Cooktop',                           'product_category': 'Service', 'charge': 110.00, 'margin': 90.00},
    'X216':  {'material_type': 'SERVICE', 'description': 'Install Gas Cooktop',                                'product_category': 'Service', 'charge': 231.00, 'margin': 34.00},
    'X222':  {'material_type': 'SERVICE', 'description': 'Install Electric Professional Range (Up to 36in)',   'product_category': 'Service', 'charge': 275.00, 'margin': 65.00},
    'X224':  {'material_type': 'SERVICE', 'description': 'Install Gas Professional Range (Up to 36in)',        'product_category': 'Service', 'charge': 275.00, 'margin': 44.00},
    'X226':  {'material_type': 'SERVICE', 'description': 'Install Microwave (Over-The-Range)',                 'product_category': 'Service', 'charge': 140.00, 'margin': 108.00},
    'X230':  {'material_type': 'SERVICE', 'description': 'Install Microwave (Countertop)',                     'product_category': 'Service', 'charge':   5.50, 'margin':  0.00},
    'X232':  {'material_type': 'SERVICE', 'description': 'Install Warming Drawer',                             'product_category': 'Service', 'charge': 140.00, 'margin': 10.00},
    'X238':  {'material_type': 'SERVICE', 'description': 'Install Electric Professional Range (>36in)',         'product_category': 'Service', 'charge': 400.00, 'margin': 53.50},
    'X251':  {'material_type': 'SERVICE', 'description': 'Install Washer',                                     'product_category': 'Service', 'charge':  18.75, 'margin':  6.25},
    'X253':  {'material_type': 'SERVICE', 'description': 'Install Electric Dryer',                             'product_category': 'Service', 'charge':  18.75, 'margin':  6.25},
    'X255':  {'material_type': 'SERVICE', 'description': 'Install Gas Dryer',                                  'product_category': 'Service', 'charge':  32.75, 'margin': 17.25},
    'X257':  {'material_type': 'SERVICE', 'description': 'Install Electric Unitized Laundry Center/Tower',     'product_category': 'Service', 'charge':  37.50, 'margin': 12.50},
    'X259':  {'material_type': 'SERVICE', 'description': 'Install Gas Unitized Laundry Center/Tower',          'product_category': 'Service', 'charge':  37.50, 'margin': 12.50},
    'X260':  {'material_type': 'SERVICE', 'description': 'Stack Laundry Unit',                                 'product_category': 'Service', 'charge':  27.50, 'margin':  0.00},
    'X261':  {'material_type': 'SERVICE', 'description': 'Install Laundry Pedestal/Riser',                     'product_category': 'Service', 'charge':  27.50, 'margin':  0.00},
    'X301':  {'material_type': 'SERVICE', 'description': 'Install Dishwasher',                                 'product_category': 'Service', 'charge':  65.50, 'margin': 34.50},
    'X303':  {'material_type': 'SERVICE', 'description': 'Dry Fit Dishwasher (No Water/Electric)',              'product_category': 'Service', 'charge':  46.75, 'margin': 28.25},
    'X328':  {'material_type': 'SERVICE', 'description': 'Install Range Vent Hood',                            'product_category': 'Service', 'charge':  65.50, 'margin': 84.50},
    'X336':  {'material_type': 'SERVICE', 'description': 'Install Duct Cover/Hood Liner/Insert (Up to 36in)',  'product_category': 'Service', 'charge':  27.50, 'margin': 72.50},
    # Codes without confirmed pricing — rates set to $0 until updated
    'X002':  {'material_type': 'SERVICE', 'description': 'Room of Choice',                                     'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X003':  {'material_type': 'SERVICE', 'description': 'Threshold',                                          'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X004':  {'material_type': 'SERVICE', 'description': 'Exterior',                                           'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X007':  {'material_type': 'SERVICE', 'description': 'Return Pickup',                                      'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X011':  {'material_type': 'SERVICE', 'description': 'Additional Labor Daily (3-Man Truck)',                'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X012':  {'material_type': 'SERVICE', 'description': 'Stair Carry (Above 3 floors)',                       'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X013':  {'material_type': 'SERVICE', 'description': 'After Hours Pickup or Delivery (Hot Shot)',           'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X014':  {'material_type': 'SERVICE', 'description': 'Military Base Delivery',                             'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X020':  {'material_type': 'SERVICE', 'description': 'Presite - per trip, not per item',                   'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X023':  {'material_type': 'SERVICE', 'description': 'Detention Time (15 minute increment)',                'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X025':  {'material_type': 'SERVICE', 'description': 'Garbage Disposal Delivered',                         'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X028':  {'material_type': 'SERVICE', 'description': 'Service Call',                                       'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X075':  {'material_type': 'SERVICE', 'description': 'Light Assembly (15 minute increment)',                'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X101':  {'material_type': 'SERVICE', 'description': 'Move Old Appliance to New Room',                     'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X102':  {'material_type': 'SERVICE', 'description': 'Large Appliance Delivery Charge (Above 600 lbs.)',   'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X210':  {'material_type': 'SERVICE', 'description': 'Install Gas Single Wall Oven',                       'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    'X998':  {'material_type': 'SERVICE', 'description': 'Daily Minimum Rate Per Route',                       'product_category': 'Service', 'charge':  0.00, 'margin': 0.00},
    # Parts that appear in the SERVICE_CODES namespace
    'WATERHOSE SS':      {'material_type': 'PART', 'description': 'Waterhose-SS (PR)',             'product_category': 'Parts', 'charge': 0.00, 'margin': 0.00},
    'LAUNDRYPACK':       {'material_type': 'PART', 'description': 'Dryer Cord, Duct & 2 Collars', 'product_category': 'Parts', 'charge': 0.00, 'margin': 0.00},
    'LAUNDRYPACK-ELECT': {'material_type': 'PART', 'description': 'Electric Laundry Pack',        'product_category': 'Parts', 'charge': 0.00, 'margin': 0.00},
    'RANGECORD':         {'material_type': 'PART', 'description': '240V Range Power Cord',        'product_category': 'Parts', 'charge': 0.00, 'margin': 0.00},
    'DW INSTALL KIT':    {'material_type': 'PART', 'description': 'Dishwasher Install Kit',       'product_category': 'Parts', 'charge': 0.00, 'margin': 0.00},
    'STEAM DRYER':       {'material_type': 'PART', 'description': 'Steam Dryer Kit',              'product_category': 'Parts', 'charge': 0.00, 'margin': 0.00},
    'STEAM DRYER ':      {'material_type': 'PART', 'description': 'Steam Dryer Kit',              'product_category': 'Parts', 'charge': 0.00, 'margin': 0.00},
    'MEMO':              {'material_type': 'PART', 'description': 'Memo',                         'product_category': 'Parts', 'charge': 0.00, 'margin': 0.00},
}

# ══════════════════════════════════════════════════════════════════
# CODE CLASSIFICATION TABLES
# ══════════════════════════════════════════════════════════════════

# Lines with these model numbers are deleted from the flat file entirely
DELETE_CODES = {
    'B003', 'B002', 'CC FEES', 'CC FEE',
    'STAIR 5-10', 'STAIR 11-15', 'STAIR 16-20', 'STAIR 21-25',
    'ACCOMM', 'MGMT ACCOMMODATION',
    'R&O-ACCTG USE ONLY', 'PENCE - ACCTG USE ONLY', 'WALSH -ACCTG USE ONLY',
    'COLAS-ACCTG USE ONLY', 'LIEN', 'CAT TAX', 'NSF', 'CHARGEBACK CC',
    'PAYROLL', 'SM WO', 'ACCTG - WARRANTIES', 'ACCOUNTING', 'REFUND',
    'WF FEES', 'TD', 'FACTORY', 'CC FEE', 'STORAGE', 'RESTOCK',
    'LATE FEE', 'SPECIAL', '99PRICEADJ', '99MISC50',
    'CUSTOMERS HOME', 'DAMAGE BY DELIVERY', 'FREIGHT',
}

# Redelivery codes — treated as X001 White Glove
REDEL_CODES = {'REDEL'}

# Model numbers that look like appliances but are actually install parts
PARTS_LOOKUP = {
    'WATERLINE':       'Refrigerator Waterline',
    'WATERHOSE SS':    'Waterhose-SS (PR)',
    'WATERHOSE RUB':   'Waterhose-Rubber (PR)',
    'GASLINE':         'Dryer/Range Gasline',
    'LAUNDRYPACK-GAS': 'Gas Flexline, Flex Duct & 2 Collars',
    'DRYERCORD':       '240V Dryer Cord',
    'DWELBOW':         'Dishwasher Elbow',
    'BRACKET':         'Dishwasher Bracket',
    'ADA DW PANS 18"': 'ADA DW Pans 18"',
    'DW PANS':         'Dishwasher Pans',
    'LAUNDRY PAN':     'Laundry Pan',
    '52525':           'Eastman Washer Pans',
    '3PRONGCORD':      'Power Cord',
    '110 POWERCORD':   'Power Cord',
    'DW KIT':          'Dishwasher Install Kit',
}

# Internal labor codes that map to HUB service X-codes
LABOR_LOOKUP = {
    'WASHERIN':             {'code': 'X251', 'description': 'Install Washer',       'material_type': 'SERVICE', 'product_category': 'Service'},
    'REFERINSTALL':         {'code': 'X151', 'description': 'Install Refrigerator', 'material_type': 'SERVICE', 'product_category': 'Service'},
    'CONVERSION-DOORSWING': {'code': 'X103', 'description': 'Reverse Door Swing',   'material_type': 'SERVICE', 'product_category': 'Service'},
}

# Internal labor codes that are third-party installs (kept as PART, not SERVICE)
LABOR_PART_CODES = {
    'TPI':            'Third Party Install',
    'CONVERSION-GAS': 'Gas Conversion (TPI)',
}

# Dryer install — resolved to gas or electric at runtime
DRYER_INSTALL_CODES = {'DRYER INSTALL'}

# Misc invoice codes to keep in the flat file, classified as PART
KEEP_AS_PART = {'RETURN', 'SALE', 'EXCH', 'ACCOMM'}

# Monday crate-status label → canonical value used in financial logic
CRATE_LABEL_MAP = {
    'IN-BOX':                  'IN-BOX',
    'OUT OF BOX':              'OUT OF BOX',
    'OUT OF BOX + INST':       'OUT OF BOX + INSTALL',
    'OUT OF BOX + INSTA':      'OUT OF BOX + INSTALL',
    'OUT OF BOX + INSTAL':     'OUT OF BOX + INSTALL',
    'OUT OF BOX + INSTALL':    'OUT OF BOX + INSTALL',
    'OUT OF BOX + INSTALL(S)': 'OUT OF BOX + INSTALL',
}

# Inventory category → HUB product category
CATEGORY_MAP = {
    'Dishwasher':                      'Dishwasher',
    'Dryer':                           'Dryer',
    'Washer':                          'Washer',
    'Range':                           'Range',
    'Rangetops':                       'Range',
    'Cooktop':                         'Range',
    'Microwave':                       'Microwave',
    'Hood':                            'Range Hood',
    'Wall Oven/Warming Drawers':       'Oven',
    'Refrigeration':                   'Refrigerator',
    'Built In Refrigeration':          'Built In Refrigeration',
    'Specialty Refrigeration':         'Refrigerator',
    'Freezer':                         'Refrigerator',
    'Appliance Accessories and Parts': 'Parts',
    'Laundry Pedestals':               'Accessories',
    'Coffee':                          'Accessories',
    'Small Appliances':                'Accessories',
    'Outdoor Appliances':              'Accessories',
    'Compact  Kitchen':                'Accessories',
    'Misc':                            'Accessories',
    'Misc Stuff':                      'Accessories',
    'Garbage Disposal':                'Accessories',
    'Water Dispensers':                'Accessories',
    'Air Quality Management':          'Accessories',
    'Dehumidifier':                    'Accessories',
    'Iron':                            'Accessories',
    'Vacuum ':                         'Accessories',
    'Trash Compactor':                 'Accessories',
    'Cookware & Utensils':             'Accessories',
    'Outdoor Furniture':               'Accessories',
}

# ══════════════════════════════════════════════════════════════════
# HUB SERVICE AREA — valid zip codes
# Orders outside this list are flagged before the flat file is built
# ══════════════════════════════════════════════════════════════════
HUB_ZIPS = {
    '97002','97003','97004','97005','97006','97007','97008','97009','97010',
    '97011','97013','97014','97015','97017','97018','97019','97022','97023',
    '97024','97026','97027','97028','97030','97031','97032','97034','97035',
    '97038','97042','97045','97048','97049','97051','97053','97054','97055',
    '97056','97060','97062','97064','97067','97068','97070','97071','97080',
    '97086','97089','97101','97106','97109','97111','97113','97114','97115',
    '97116','97117','97119','97123','97124','97125','97127','97128','97132',
    '97133','97137','97140','97144','97148','97201','97202','97203','97204',
    '97205','97206','97209','97210','97211','97212','97213','97214','97215',
    '97216','97217','97218','97219','97220','97221','97222','97223','97224',
    '97225','97227','97229','97230','97231','97232','97233','97236','97239',
    '97266','97267','97281','97301','97302','97304','97305','97306','97310',
    '97317','97321','97322','97325','97330','97331','97333','97338','97351',
    '97352','97360','97361','97362','97370','97371','97375','97381','97385',
    '97389','97392','97448',
    '98601','98603','98604','98605','98606','98607','98610','98616','98625',
    '98626','98629','98632','98642','98648','98651','98660','98661','98662',
    '98663','98664','98665','98666','98667','98668','98671','98672','98674',
    '98675','98682','98683','98684','98685','98686','98687',
}

# Flat-file column order (must match HUB template sheet header row)
FLAT_COLS = [
    'location_code', 'order_number', 'delivery_id', 'delivery_date',
    'type_of_order', 'delivery_name', 'delivery_phone_1', 'delivery_phone_2',
    'delivery_address_1', 'delivery_address_2', 'delivery_unit', 'delivery_floor',
    'delivery_city', 'delivery_state', 'delivery_postal_code', 'reference_number',
    'material_type', 'item_code', 'item_quantity', 'item_material_description',
    'item_product_category', 'item_points', 'item_weight', 'service_haul_away',
]


# ══════════════════════════════════════════════════════════════════
# MILEAGE — Open Route Service geocode + driving distance
# ══════════════════════════════════════════════════════════════════

def geocode_address(address, api_key):
    """Return [lon, lat] for *address* using the ORS geocoding API."""
    url = "https://api.openrouteservice.org/geocode/search"
    try:
        r = requests.get(url, params={'api_key': api_key, 'text': address, 'size': 1}, timeout=10)
        features = r.json().get('features', [])
        if features:
            return features[0]['geometry']['coordinates']
    except Exception as exc:
        print(f"  ⚠️  Geocode failed for '{address}': {exc}")
    return None


def get_drive_miles(origin_coords, dest_coords, api_key):
    """Return driving distance in miles between two [lon, lat] points."""
    url = "https://api.openrouteservice.org/v2/directions/driving-car"
    try:
        r = requests.post(
            url,
            json={'coordinates': [origin_coords, dest_coords]},
            headers={'Authorization': api_key},
            timeout=10,
        )
        meters = r.json()['routes'][0]['summary']['distance']
        return round(meters / 1609.34, 1)
    except Exception as exc:
        print(f"  ⚠️  Distance lookup failed: {exc}")
    return None


# ══════════════════════════════════════════════════════════════════
# MONDAY.COM — Delivery Scheduler crate-status lookup
# ══════════════════════════════════════════════════════════════════

def get_crate_status(token, board_id, order_numbers, delivery_date,
                     crate_col_id, date_col_id):
    """Query the Monday Delivery Scheduler for crate status per order.

    Parameters
    ----------
    token          : Monday API token
    board_id       : int, Monday board ID
    order_numbers  : iterable of base order number strings
    delivery_date  : datetime, today's delivery date
    crate_col_id   : str, Monday column ID for the crate-status dropdown
    date_col_id    : str, Monday column ID for the requested delivery date

    Returns
    -------
    result : dict  {order_num: 'IN-BOX' | 'OUT OF BOX' | 'OUT OF BOX + INSTALL'}
    flags  : list  of {'order': str, 'issue': str} dicts for orders with missing status
    """
    if not token:
        print("  ⚠️  MONDAY_API_TOKEN not set — crate charges skipped")
        return {}, []

    target_orders  = {str(o).split('-')[0] for o in order_numbers}
    target_date_md = delivery_date.strftime('%Y-%m-%d')
    result         = {}
    on_board_today = set()
    flags          = []
    headers        = {"Authorization": token, "Content-Type": "application/json"}
    cursor         = None

    print("\n  Fetching crate status from Monday Delivery Scheduler...")

    while True:
        col_ids = f'"{crate_col_id}", "{date_col_id}"'
        if cursor:
            query = (
                """query ($boardId: [ID!], $cursor: String!) {
                  boards(ids: $boardId) {
                    items_page(limit: 500, cursor: $cursor) {
                      cursor
                      items { name column_values(ids: [%s]) { id text value } }
                    }
                  }
                }""" % col_ids
            )
            variables = {"boardId": str(board_id), "cursor": cursor}
        else:
            query = (
                """query ($boardId: [ID!]) {
                  boards(ids: $boardId) {
                    items_page(limit: 500) {
                      cursor
                      items { name column_values(ids: [%s]) { id text value } }
                    }
                  }
                }""" % col_ids
            )
            variables = {"boardId": str(board_id)}

        try:
            r    = requests.post("https://api.monday.com/v2",
                                 json={"query": query, "variables": variables},
                                 headers=headers, timeout=15)
            data = r.json()
            if "errors" in data:
                print(f"  ⚠️  Monday API error: {data['errors']} — crate charges skipped")
                return {}, []
            page  = data["data"]["boards"][0]["items_page"]
            items = page["items"]
        except Exception as exc:
            print(f"  ⚠️  Monday API call failed: {exc} — crate charges skipped")
            return {}, []

        for item in items:
            raw_name   = item["name"].strip()
            base_order = raw_name.split()[0] if raw_name else ''
            if base_order not in target_orders:
                continue

            col_map    = {c["id"]: c for c in item.get("column_values", [])}
            date_value = (col_map.get(date_col_id) or {}).get("value") or ""
            try:
                item_date = json.loads(date_value).get("date", "") if date_value else ""
            except Exception:
                item_date = ""
            if item_date != target_date_md:
                continue

            on_board_today.add(base_order)
            label_raw = ((col_map.get(crate_col_id) or {}).get("text") or "").strip().upper()
            canonical  = CRATE_LABEL_MAP.get(label_raw)
            if canonical:
                result[base_order] = canonical
            elif label_raw:
                print(f"  ⚠️  Unrecognized crate label for order {base_order}: '{label_raw}'")

        cursor = page.get("cursor")
        if not cursor:
            break

    # Orders on the board today but crate status left blank
    for o in sorted(on_board_today - result.keys()):
        msg = f"Order {o}: on Delivery Scheduler for today but crate status is blank"
        print(f"  ⚠️  {msg} — crate charge skipped")
        flags.append({'order': o, 'issue': msg})

    # Orders not on the board for today's date at all
    for o in sorted(target_orders - on_board_today):
        msg = f"Order {o}: not found on Delivery Scheduler for {delivery_date.strftime('%m/%d/%Y')}"
        print(f"  ⚠️  {msg} — crate charge skipped")
        flags.append({'order': o, 'issue': msg})

    print(f"  ✅ Crate status fetched: {len(result)} orders")
    for o, s in sorted(result.items()):
        print(f"     Order {o}: {s}")

    return result, flags


# ══════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════

def get_pricing(model_key, qty):
    """Return (charge, margin) for a service code × quantity."""
    if model_key in SERVICE_CODES:
        info = SERVICE_CODES[model_key]
        return round(info['charge'] * qty, 2), round(info['margin'] * qty, 2)
    return 0.00, 0.00


def has_model_lines(order_rows):
    """Return True if the order contains at least one physical appliance line."""
    for _, row in order_rows.iterrows():
        model_key = str(row.get('Model Number', '')).strip().upper()
        if model_key and model_key not in SERVICE_CODES:
            return True
    return False


def get_dryer_xcode(order_group, inv_lookup):
    """Resolve 'DRYER INSTALL' to gas (X255) or electric (X253) by checking inventory."""
    for _, row in order_group.iterrows():
        model_key = str(row.get('Model Number', '')).strip().upper()
        if model_key in inv_lookup:
            desc = str(inv_lookup[model_key].get('description', '')).upper()
            if any(word in desc for word in ('GAS', 'NATURAL GAS', 'LP')):
                return 'X255', 'Install Gas Dryer'
    return 'X253', 'Install Electric Dryer'


def validate_x001(order_rows, order_num, flags):
    """Ensure exactly one X001 (White Glove) line exists per order."""
    rows      = order_rows.copy().reset_index(drop=True)
    x001_mask = rows['Model Number'].astype(str).str.upper().str.strip() == 'X001'
    count     = x001_mask.sum()
    if count == 0:
        new_row                 = rows.iloc[0].copy()
        new_row['Model Number'] = 'X001'
        new_row['Description']  = 'White Glove'
        new_row['Qty']          = 1
        rows = pd.concat([rows, new_row.to_frame().T], ignore_index=True)
        flags.append({'order': order_num, 'issue': '➕ X001 missing — auto-added (White Glove $80.00)'})
    elif count > 1:
        extra_idx = rows[x001_mask].index[1:]
        rows      = rows.drop(extra_idx).reset_index(drop=True)
        flags.append({'order': order_num, 'issue': f'🔁 {count} X001 lines — kept first, removed {len(extra_idx)} duplicate(s)'})
    return rows


def split_rows_across_trucks(order_rows, num_trucks):
    """Distribute order rows evenly across *num_trucks* truck sub-orders."""
    trucks = [[] for _ in range(num_trucks)]
    for _, row in order_rows.iterrows():
        try:
            qty = int(float(row.get('Qty', 1)))
        except Exception:
            qty = 1
        base = qty // num_trucks
        rem  = qty % num_trucks
        for t in range(num_trucks):
            r       = row.copy()
            r['Qty'] = base + (rem if t == num_trucks - 1 else 0)
            trucks[t].append(r)
    return [pd.DataFrame(t) for t in trucks]


def order_total_charge(order_rows, mileage_amt, fuel_amt):
    """Sum all service charges for an order (used to determine truck splits)."""
    total = mileage_amt + fuel_amt
    for _, row in order_rows.iterrows():
        model_key = str(row.get('Model Number', '')).strip().upper()
        try:
            qty = int(float(row.get('Qty', 1)))
        except Exception:
            qty = 1
        charge, _ = get_pricing(model_key, qty)
        total += charge
    return max(total, MIN_CHARGE)


def prompt_int(prompt_text, allow_zero=True):
    """Prompt the user for an integer, re-asking on invalid input."""
    while True:
        try:
            val = int(input(prompt_text).strip())
            if not allow_zero and val < 1:
                print("  Please enter a number greater than 0.")
                continue
            return val
        except ValueError:
            print("  Please enter a whole number.")


def make_row(formatted_order, delivery_date, row, ref_num, material_type,
             item_code, qty, description, category, haul_away,
             line_charge, line_margin, running_charge, running_margin,
             location_code, floor_num=''):
    """Build one output row dict for the HUB flat file."""
    phone1  = str(row.get('Shipping Phone',   '')).strip() if pd.notna(row.get('Shipping Phone'))   else ''
    phone2  = str(row.get('Shipping Phone 2', '')).strip() if pd.notna(row.get('Shipping Phone 2')) else ''
    zip_raw = str(row.get('Shipping Zip', '')).strip()
    return {
        'location_code':             location_code,
        'order_number':              formatted_order,
        'delivery_id':               formatted_order,
        'delivery_date':             delivery_date,
        'type_of_order':             ('Service'
                                      if str(row.get('Model Number', '')).strip().upper() == 'MEMO'
                                      else str(row.get('Delivery/Pick-up Type', 'Delivery'))),
        'delivery_name':             str(row.get('Shipping Customer', '')).strip(),
        'delivery_phone_1':          phone1,
        'delivery_phone_2':          phone2,
        'delivery_address_1':        str(row.get('Shipping Address', '')).strip(),
        'delivery_address_2':        '',
        'delivery_unit':             '',
        'delivery_floor':            str(floor_num) if floor_num else '',
        'delivery_city':             str(row.get('Shipping City', '')).strip(),
        'delivery_state':            str(row.get('Shipping State', '')).strip(),
        'delivery_postal_code':      (int(zip_raw[:5])
                                      if len(zip_raw) >= 5 and zip_raw[:5].isdigit()
                                      else zip_raw),
        'reference_number':          ref_num,
        'material_type':             material_type,
        'item_code':                 item_code,
        'item_quantity':             qty,
        'item_material_description': description,
        'item_product_category':     category,
        'item_points':               '',
        'item_weight':               '',
        'service_haul_away':         haul_away,
        '_line_charge':              line_charge,
        '_running_charge':           round(running_charge, 2),
        '_line_margin':              line_margin,
        '_running_margin':           round(running_margin, 2),
    }


def clean_dollar(val):
    """Parse a dollar string such as '$1,234.56' into a float."""
    if pd.isna(val):
        return 0.0
    s = str(val).replace('$', '').replace(',', '').strip()
    return float(s) if s not in ('', 'nan') else 0.0


def clean_pct(val):
    """Parse a percentage string such as '12.5%' into a decimal float."""
    if pd.isna(val):
        return None
    s = str(val).replace('%', '').strip()
    try:
        return float(s) / 100
    except ValueError:
        return None


def get_serial_units(order_num, model, df_serial_alloc, delivery_date):
    """Return [(cost, inventory_id), ...] for units allocated to this order/model/date.

    Only units whose Est. Delivery Date matches *delivery_date* are returned;
    ASAP or blank units belong to future deliveries.
    """
    if df_serial_alloc is None:
        return []
    base_order = str(order_num).split('-')[0]
    today_str  = f"{delivery_date.month}/{delivery_date.day}/{delivery_date.year}"
    matches = df_serial_alloc[
        (df_serial_alloc['Order #'] == base_order) &
        (df_serial_alloc['Model']   == model.upper().strip()) &
        (df_serial_alloc['Est. Delivery Date'].astype(str).str.strip() == today_str)
    ]
    return [
        (
            round(row['Cost'], 2),
            (str(int(float(row['Inventory Id'])))
             if str(row['Inventory Id']).replace('.', '').isdigit()
             else str(row['Inventory Id'])),
        )
        for _, row in matches.iterrows()
    ]


def write_financial_sheet(ws, df, sum_cols, avg_cols, dollar_cols, header_fills):
    """Write a DataFrame to an openpyxl worksheet with totals row and formatting."""
    bold_font   = Font(bold=True)
    normal_font = Font(bold=False)
    dollar_fmt  = '_("$"* #,##0.00_);_("$"* (  #,##0.00);_("$"* "-"??_);_(@_)'
    pct_fmt     = '0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )

    col_names = list(df.columns)
    for col_idx, col_name in enumerate(col_names, 1):
        cell        = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = bold_font
        cell.border = thin_border
        fill_rgb    = header_fills.get(col_name)
        if fill_rgb:
            cell.fill = PatternFill(fill_type='solid', fgColor=fill_rgb)
        if col_name in dollar_cols:
            cell.number_format = dollar_fmt

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(col_names, 1):
            cell       = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.font  = normal_font
            if col_name in dollar_cols:
                cell.number_format = dollar_fmt
            elif col_name in avg_cols:
                cell.number_format = pct_fmt

    last_data_row = len(df) + 1
    total_row     = last_data_row + 3
    ws.cell(row=total_row, column=1, value='total').font = bold_font

    for col_idx, col_name in enumerate(col_names, 1):
        col_letter = get_column_letter(col_idx)
        cell       = ws.cell(row=total_row, column=col_idx)
        cell.font  = bold_font
        fill_rgb   = header_fills.get(col_name)
        if col_name in sum_cols:
            cell.value         = f'=SUM({col_letter}2:{col_letter}{last_data_row})'
            cell.number_format = dollar_fmt
            if fill_rgb:
                cell.fill = PatternFill(fill_type='solid', fgColor=fill_rgb)
        elif col_name in avg_cols:
            # Blended margin = profit / sale, not average of per-row percentages
            profit_col = get_column_letter(col_idx - 1)
            _sale_names = {'sale', 'total_sale', 'svc_sale', 'prd_sale'}
            _sale_idx   = next(
                (ci + 1 for ci, cn in reversed(list(enumerate(col_names[:col_idx - 1])))
                 if cn in _sale_names),
                col_idx - 2,
            )
            sale_col           = get_column_letter(_sale_idx)
            cell.value         = f'=IFERROR({profit_col}{total_row}/{sale_col}{total_row}*100,0)'
            cell.number_format = pct_fmt
            if fill_rgb:
                cell.fill = PatternFill(fill_type='solid', fgColor=fill_rgb)


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    import glob as _glob

    prefix    = CONFIG['output_prefix']
    loc_code  = CONFIG['location_code']
    wh_addr   = CONFIG['warehouse_address']
    board_id  = CONFIG['delivery_scheduler_board_id']
    crate_col = CONFIG['monday_crate_col_id']
    date_col  = CONFIG['monday_date_col_id']

    # ── Resolve environment ────────────────────────────────────────
    if _IS_COLAB:
        from google.colab import userdata as _userdata
        ors_key    = _userdata.get('ORS_API_KEY')
        mon_token  = _userdata.get('MONDAY_API_TOKEN')
        content_dir = '/content'
        save_dir    = '.'
    else:
        from dotenv import load_dotenv
        load_dotenv()
        ors_key    = os.getenv('ORS_API_KEY', '')
        mon_token  = os.getenv('MONDAY_API_TOKEN', '')
        content_dir = os.getenv('CONTENT_DIR', CONFIG['local_content_dir'])
        export_dir  = os.getenv('EXPORT_DIR',  CONFIG['local_export_dir'])
        save_dir    = export_dir  # resolved per-date below

    # ══════════════════════════════════════════════════════════════
    # STEP 1 — LOAD FILES
    # ══════════════════════════════════════════════════════════════
    print('=' * 60)
    print(f'  HUB FLAT FILE GENERATOR  v5')
    print('=' * 60)

    batch_files = sorted(
        [f for f in _glob.glob(f'{content_dir}/*.xlsx') if 'bulk-invoice' in f.lower()],
        key=len,
    )
    csv_files   = [f for f in _glob.glob(f'{content_dir}/*.csv') if 'model-inventory' in f.lower()]

    if not batch_files or not csv_files:
        print('  Upload all files: batch invoice (.xlsx), inventory (.csv), orders detail (.csv)')
        from google.colab import files as _files
        uploaded = _files.upload()
        for fname, fdata in uploaded.items():
            with open(f'{content_dir}/{fname}', 'wb') as fh:
                fh.write(fdata)
        batch_files = sorted(
            [f for f in _glob.glob(f'{content_dir}/*.xlsx') if 'FlatFile' not in f], key=len
        )
        csv_files = [f for f in _glob.glob(f'{content_dir}/*.csv') if 'orders-detail' not in f]

    if not batch_files:
        raise FileNotFoundError('No batch invoice found')
    df_raw = pd.read_excel(batch_files[0])
    print(f'  Batch invoice : {batch_files[0]}')
    print(f'                  {len(df_raw)} rows, {df_raw["Order #"].nunique()} orders')

    if not csv_files:
        raise FileNotFoundError('No inventory CSV found')
    df_inv = pd.read_csv(csv_files[0])
    print(f'  Inventory CSV : {csv_files[0]}')

    orders_files = [f for f in _glob.glob(f'{content_dir}/*.csv') if 'orders-detail' in f]
    if orders_files:
        df_orders = pd.read_csv(orders_files[0])
        df_orders['Cost_clean']      = df_orders['Cost'].apply(clean_dollar)
        df_orders['SalePrice_clean'] = df_orders['SalePrice'].apply(clean_dollar)
        df_orders['Margin_clean']    = df_orders['Margin'].apply(clean_pct)
        df_orders['ExtPrice_clean']  = df_orders['Ext. Price'].apply(clean_dollar)
        df_orders['Order #']         = df_orders['Order #'].astype(str)
        print(f'  Orders detail : {orders_files[0]}')
        print(f'                  {len(df_orders)} lines, {df_orders["Order #"].nunique()} orders (before date filter)')
        has_orders_detail = True
    else:
        print('  No orders detail CSV — financial report will be skipped')
        df_orders         = None
        has_orders_detail = False

    serial_files = [f for f in _glob.glob(f'{content_dir}/*.csv') if 'serial-number-inventory' in f]
    if serial_files:
        df_serial       = pd.read_csv(serial_files[0])
        df_serial['Order #'] = df_serial['Order #'].astype(str).str.strip()
        df_serial['Model']   = df_serial['Model'].astype(str).str.strip().str.upper()
        df_serial['Cost']    = df_serial['Cost'].apply(clean_dollar)
        df_serial_alloc = df_serial[
            df_serial['Order #'].notna() &
            ~df_serial['Order #'].isin({'', 'nan', 'NaN'}) &
            df_serial['Order #'].str.match(r'^\d+\.?\d*$', na=False)
        ].copy()
        df_serial_alloc['Order #'] = df_serial_alloc['Order #'].str.replace('.0', '', regex=False)
        print(f'  Serial inv.   : {serial_files[0]}')
        print(f'                  {len(df_serial)} units, {len(df_serial_alloc)} allocated')
    else:
        print('  No serial inventory CSV — using standard costs')
        df_serial_alloc = None

    # ══════════════════════════════════════════════════════════════
    # STEP 2 — INVENTORY LOOKUP
    # ══════════════════════════════════════════════════════════════
    inv_lookup = {}
    for _, row in df_inv.iterrows():
        model = str(row['Model']).upper()
        if not model:
            continue
        cat = row['Category']
        inv_lookup[model] = {
            'material_type':    'PART' if cat in PART_CATEGORIES else 'MODEL',
            'description':      row['Description'],
            'product_category': CATEGORY_MAP.get(cat, 'Accessories'),
        }
    print(f'  ✅ Inventory lookup: {len(inv_lookup)} entries')

    # ══════════════════════════════════════════════════════════════
    # STEP 3 — PARSE DATE
    # ══════════════════════════════════════════════════════════════
    sample_date = df_raw['ShipDate'].iloc[0]
    try:
        delivery_date = datetime.strptime(str(sample_date), '%m/%d/%Y')
    except ValueError:
        delivery_date = pd.to_datetime(sample_date).to_pydatetime()
    date_str = delivery_date.strftime('%m%d%y')
    print(f'\n  📅 Delivery date: {delivery_date.strftime("%B %d, %Y")}')

    if has_orders_detail:
        date_filter   = delivery_date.strftime('%m/%d/%Y')
        before        = len(df_orders)
        df_orders     = df_orders[
            df_orders['Est. Delivery'].astype(str).str.strip() == date_filter
        ].copy()
        print(f'  📋 Orders detail filtered: {before} → {len(df_orders)} lines matching {date_filter}')

    # Local save path (year / month subfolders)
    if not _IS_COLAB:
        save_dir = os.path.join(
            export_dir,
            delivery_date.strftime('%Y'),
            delivery_date.strftime('%B').upper(),
        )
        os.makedirs(save_dir, exist_ok=True)
        print(f'  Save folder: {save_dir}')

    # ══════════════════════════════════════════════════════════════
    # STEP 4 — IDENTIFY MULTIFAMILY ORDERS
    # ══════════════════════════════════════════════════════════════
    mf_order_set = set(
        df_raw[df_raw['Billing Customer Type'].str.contains('MULTI', na=False)]['Order #'].astype(str)
    )
    print(f'  🏢 Multifamily orders: {mf_order_set or "None"}')

    # ══════════════════════════════════════════════════════════════
    # STEP 5 — MILEAGE (Open Route Service)
    # ══════════════════════════════════════════════════════════════
    print('\n' + '─' * 60)
    print('  📍 CALCULATING MILEAGE (Open Route Service)')
    print(f'  Warehouse: {wh_addr}')
    print('─' * 60)

    warehouse_coords = geocode_address(wh_addr, ors_key)
    if not warehouse_coords:
        print('  ⚠️  Could not geocode warehouse — mileage charges will be skipped')

    order_mileage        = {}
    order_mileage_charge = {}
    order_fuel_charge    = {}
    over_mileage_flags   = []
    all_orders           = df_raw['Order #'].astype(str).unique()

    for order_num in all_orders:
        match     = df_raw[df_raw['Order #'].astype(str) == order_num].iloc[0]
        address   = (f"{match['Shipping Address']}, {match['Shipping City']}, "
                     f"{match['Shipping State']} {match['Shipping Zip']}")
        stop_name = match['Shipping Customer']
        miles     = None

        if warehouse_coords:
            dest_coords = geocode_address(address, ors_key)
            time.sleep(0.3)
            if dest_coords:
                miles = get_drive_miles(warehouse_coords, dest_coords, ors_key)
                time.sleep(0.3)
                if miles and miles > 200:
                    miles = None  # discard implausible geocode result

        if miles is None:
            order_mileage[order_num]        = None
            order_mileage_charge[order_num] = 0.00
            order_fuel_charge[order_num]    = 0.00
            print(f'  ⚠️  Order {order_num} ({stop_name}): mileage lookup failed — $0')
        elif miles < MILEAGE_FREE_MAX:
            order_mileage[order_num]        = miles
            order_mileage_charge[order_num] = 0.00
            order_fuel_charge[order_num]    = FUEL_SHORT
            print(f'  ✅ Order {order_num} ({stop_name}): {miles} mi — fuel ${FUEL_SHORT}')
        elif miles <= MILEAGE_MAX:
            order_mileage[order_num]        = miles
            order_mileage_charge[order_num] = MILEAGE_CHARGE
            order_fuel_charge[order_num]    = FUEL_LONG
            print(f'  ✅ Order {order_num} ({stop_name}): {miles} mi — fuel ${FUEL_LONG} + mileage ${MILEAGE_CHARGE:.0f}')
        else:
            order_mileage[order_num]        = miles
            order_mileage_charge[order_num] = 0.00
            order_fuel_charge[order_num]    = FUEL_LONG
            over_mileage_flags.append(order_num)
            print(f'  🚛 Order {order_num} ({stop_name}): {miles} mi — OVER {MILEAGE_MAX} mi, flagged')

    # ══════════════════════════════════════════════════════════════
    # STEP 5.5 — MONDAY CRATE STATUS
    # ══════════════════════════════════════════════════════════════
    order_crate_status, crate_flags = get_crate_status(
        mon_token, board_id, all_orders, delivery_date, crate_col, date_col
    )

    # ══════════════════════════════════════════════════════════════
    # STEP 6 — MULTIFAMILY PROMPTS (units + floor charges)
    # ══════════════════════════════════════════════════════════════
    order_units       = {}
    order_floor_units = {}

    if mf_order_set:
        print('\n' + '─' * 60)
        print('  🏢 MULTIFAMILY ORDER DETAILS')
        print('─' * 60)
        for order_num in sorted(mf_order_set):
            match       = df_raw[df_raw['Order #'].astype(str) == order_num]
            stop_name   = match['Shipping Customer'].iloc[0] if len(match) else order_num
            city        = match['Shipping City'].iloc[0]     if len(match) else ''
            order_group = match
            is_delivery = has_model_lines(order_group)

            print(f'\n  Order {order_num} — {stop_name}, {city}')
            print(f'  Mode: {"DELIVERY (product + install)" if is_delivery else "INSTALL ONLY (product on site)"}')

            units                      = prompt_int('  How many units being serviced today? ', allow_zero=False)
            order_units[order_num]     = units
            order_floor_units[order_num] = 0

            if is_delivery:
                if input('  Any units on floor 4 or above? (y/n) ').strip().lower() == 'y':
                    num_floors  = prompt_int('  How many floors total above floor 3? ', allow_zero=False)
                    floor_total = 0.0
                    for f_idx in range(num_floors):
                        floor_num  = 4 + f_idx
                        multiplier = floor_num - 3
                        pieces     = prompt_int(f'  Floor {floor_num} — how many pieces? ')
                        charge     = round(pieces * FLOOR_CHARGE * multiplier, 2)
                        floor_total += charge
                        if pieces > 0:
                            print(f'  → Floor {floor_num}: {pieces} × ${FLOOR_CHARGE} × {multiplier} = ${charge:.2f}')
                    order_floor_units[order_num] = floor_total
                    print(f'  → Total floor charge: ${floor_total:.2f}')
                else:
                    print('  → No floor charge')
            else:
                print(f'  → Install only: no floor charge, X018 threshold applies ({units} units × $33.00)')

    # ══════════════════════════════════════════════════════════════
    # STEP 7 — BUILD FLAT FILE
    # ══════════════════════════════════════════════════════════════
    output_rows      = []
    errors           = []
    x001_flags       = []
    out_of_area      = []
    zero_price_codes = set()

    for order_num_raw, order_group in df_raw.groupby('Order #', sort=False):
        order_num   = str(order_num_raw).strip()
        is_mf       = order_num in mf_order_set
        is_delivery = has_model_lines(order_group)
        mileage_amt = order_mileage_charge.get(order_num, 0.00)
        fuel_amt    = order_fuel_charge.get(order_num, 0.00)

        # Normalize variant codes before validation
        order_group = order_group.copy()
        order_group['Model Number'] = order_group['Model Number'].apply(
            lambda m: 'X001' if str(m).strip().upper() in REDEL_CODES else m
        )
        order_group['Model Number'] = order_group['Model Number'].apply(
            lambda m: 'X103' if str(m).strip().upper() in {'X103A', 'X103B'} else m
        )
        order_group = validate_x001(order_group, order_num, x001_flags)

        units       = order_units.get(order_num, 1)
        floor_units = order_floor_units.get(order_num, 0)

        if is_mf:
            total_charge = order_total_charge(order_group, mileage_amt, fuel_amt)
            num_trucks   = max(1, math.ceil(total_charge / MIN_CHARGE))
            truck_groups = split_rows_across_trucks(order_group, num_trucks)
        else:
            num_trucks   = 1
            truck_groups = [order_group]

        for truck_seq, truck_rows in enumerate(truck_groups, 1):
            formatted_order = f'{order_num}-{date_str}-{truck_seq}'
            ref_num         = 0
            running_charge  = 0.0
            running_margin  = 0.0

            # Mileage + fuel tracked in running totals (not written to flat file)
            running_charge += mileage_amt + fuel_amt

            # X018 threshold (multifamily only) — running total only
            if is_mf and units > 0:
                running_charge += round(33.00 * units, 2)

            # Floor charge — running total only
            if is_delivery and floor_units > 0:
                running_charge += floor_units

            for _, row in truck_rows.iterrows():
                model_raw = str(row.get('Model Number', '')).strip()
                model_key = model_raw.upper()
                ref_num  += 1

                try:
                    qty = int(float(row.get('Qty', 1)))
                except Exception:
                    qty = 1

                if model_key in DELETE_CODES:
                    ref_num -= 1
                    continue

                if qty < 0:
                    output_rows.append(make_row(
                        formatted_order, delivery_date, row, ref_num,
                        'PART', 'MEMO', 1,
                        f'RMA Parts - {model_raw}', 'Parts', 'N',
                        0.00, 0.00, running_charge, running_margin, loc_code,
                    ))
                    continue

                is_memo = model_key == 'MEMO'

                if model_key in SERVICE_CODES:
                    info             = SERVICE_CODES[model_key]
                    material_type    = info['material_type']
                    item_description = (str(row.get('Description', '')).strip()
                                        if is_memo else info['description'])
                    product_category = info['product_category']
                    line_charge, line_margin = get_pricing(model_key, qty)
                    if info['charge'] == 0.00 and info['material_type'] == 'SERVICE':
                        zero_price_codes.add(model_key)

                elif model_key in PARTS_LOOKUP:
                    material_type    = 'PART'
                    item_description = PARTS_LOOKUP[model_key]
                    product_category = 'Parts'
                    line_charge, line_margin = 0.00, 0.00

                elif model_key in LABOR_PART_CODES:
                    material_type    = 'PART'
                    item_description = LABOR_PART_CODES[model_key]
                    product_category = 'Parts'
                    line_charge, line_margin = 0.00, 0.00

                elif model_key in DRYER_INSTALL_CODES:
                    xcode, xdesc     = get_dryer_xcode(order_group, inv_lookup)
                    material_type    = 'SERVICE'
                    model_raw        = xcode
                    item_description = xdesc
                    product_category = 'Service'
                    line_charge, line_margin = get_pricing(xcode, qty)

                elif model_key in LABOR_LOOKUP:
                    info             = LABOR_LOOKUP[model_key]
                    material_type    = info['material_type']
                    model_raw        = info['code']
                    item_description = info['description']
                    product_category = info['product_category']
                    line_charge, line_margin = get_pricing(info['code'], qty)

                elif model_key in KEEP_AS_PART:
                    material_type    = 'PART'
                    item_description = str(row.get('Description', '')).strip()
                    product_category = 'Parts'
                    line_charge, line_margin = 0.00, 0.00

                elif model_key in inv_lookup:
                    info             = inv_lookup[model_key]
                    material_type    = info['material_type']
                    item_description = info['description']
                    product_category = info['product_category']
                    line_charge, line_margin = 0.00, 0.00

                else:
                    material_type    = 'MODEL'
                    item_description = str(row.get('Description', '')).strip()
                    product_category = ''
                    line_charge, line_margin = 0.00, 0.00
                    errors.append({'order': order_num, 'model': model_raw,
                                   'issue': '⚠️  Not in inventory or service code table'})

                item_description = str(item_description)
                if len(item_description) > 50 and not is_memo:
                    errors.append({'order': order_num, 'model': model_raw,
                                   'issue': f'✂️  Description truncated: "{item_description[:50]}"'})

                running_charge += line_charge
                running_margin += line_margin
                haul_away       = 'Y' if str(row.get('Description', '')).upper().strip().startswith('HAUL') else 'N'
                desc_for_row    = item_description if is_memo else item_description[:50]

                output_rows.append(make_row(
                    formatted_order, delivery_date, row, ref_num,
                    material_type, model_raw, qty, desc_for_row,
                    product_category, haul_away,
                    line_charge, line_margin, running_charge, running_margin, loc_code,
                ))

    df_output = pd.DataFrame(output_rows)
    print(f'\n  ✅ Flat file built: {len(df_output)} rows, {df_output["order_number"].nunique()} orders')

    # ══════════════════════════════════════════════════════════════
    # STEP 8 — STOP SUMMARY
    # ══════════════════════════════════════════════════════════════
    print('\n' + '=' * 60)
    print('  STOP SUMMARY — CHARGES & MARGIN')
    print('=' * 60)

    summary = (
        df_output.sort_values('reference_number')
        .groupby('order_number')
        .agg(
            name         = ('delivery_name',   'first'),
            city         = ('delivery_city',   'first'),
            total_lines  = ('item_code',       'count'),
            total_charge = ('_running_charge', 'last'),
            total_margin = ('_running_margin', 'last'),
        )
        .reset_index()
    )

    print(f"\n{'Order':<22} {'Name':<24} {'Lines':>5} {'Charge':>9} {'Margin':>9}  {'Mi':>5}")
    print('-' * 75)
    for _, r in summary.iterrows():
        order_key = str(r['order_number']).split('-')[0]
        mi        = order_mileage.get(order_key)
        mi_str    = str(mi) if mi is not None else '-'
        print(f"{str(r['order_number']):<22} {str(r['name'])[:23]:<24} "
              f"{r['total_lines']:>5}  ${r['total_charge']:>8.2f}  ${r['total_margin']:>8.2f}  {mi_str:>5}")

    grand_charge = summary['total_charge'].sum()
    grand_margin = summary['total_margin'].sum()
    total_fuel   = sum(order_fuel_charge.values())
    total_stops  = df_output['order_number'].nunique()

    print('\n' + '─' * 75)
    while True:
        try:
            hub_trucks = int(input('  How many HUB trucks are running today? ').strip())
            if hub_trucks >= 1:
                break
            print('  Please enter at least 1.')
        except ValueError:
            print('  Please enter a whole number.')

    truck_minimum = hub_trucks * MIN_CHARGE
    shortfall     = max(0, truck_minimum - grand_charge)
    avg_per_stop  = grand_charge / total_stops if total_stops > 0 else 0
    stops_needed  = math.ceil(shortfall / avg_per_stop) if avg_per_stop > 0 and shortfall > 0 else 0

    print(f"\n{'─' * 75}")
    print(f"  {'DAY TOTAL CHARGES':<42}  ${grand_charge:>8,.2f}")
    print(f"  {'DAY TOTAL MARGIN':<42}  ${grand_margin:>8,.2f}")
    print(f"  {'FUEL SURCHARGES INCLUDED':<42}  ${total_fuel:>8,.2f}")
    print(f"\n  ${MIN_CHARGE:,.0f} MINIMUM CHECK:")
    print(f"  Running {hub_trucks} truck(s) → need ${truck_minimum:,.2f}")
    print(f"  Current charges: ${grand_charge:,.2f}  ({total_stops} stops)")
    if shortfall == 0:
        print(f"  ✅ Charges cover {hub_trucks} truck minimum(s)")
        print(f"  Surplus: ${grand_charge - truck_minimum:,.2f} above minimum")
    else:
        print(f"  ⚠️  SHORT ${shortfall:,.2f} to cover {hub_trucks} truck(s)")
        print(f"  Avg charge/stop: ${avg_per_stop:,.2f} → need ~{stops_needed} more stop(s)")
        prev_min = (hub_trucks - 1) * MIN_CHARGE
        covered  = '✅ covered' if grand_charge >= prev_min else f'still short ${prev_min - grand_charge:,.2f}'
        print(f"  Or reduce to {hub_trucks - 1} truck(s) — need ${prev_min:,.2f} ({covered})")

    hub_max = 9 * hub_trucks
    if total_stops > hub_max:
        print(f"\n  🚛 STOP COUNT WARNING: {total_stops} stops exceeds HUB max of {hub_max} ({hub_trucks} trucks × 9)")
    elif total_stops == hub_max:
        print(f"\n  ⚠️  At HUB maximum: {total_stops}/{hub_max} stops")
    else:
        print(f"\n  ✅ Stop count: {total_stops}/{hub_max} HUB max ({hub_trucks} trucks)")

    # ══════════════════════════════════════════════════════════════
    # STEP 9 — FLAGS
    # ══════════════════════════════════════════════════════════════
    if out_of_area:
        print(f"\n{'=' * 60}")
        print(f"  🚫 OUT OF HUB SERVICE AREA ({len(out_of_area)} order(s))")
        print('=' * 60)
        for o in out_of_area:
            print(f"  Order {o['order']} | {o['name']} | {o['zip']} — {o['city']}, {o['state']}")
        print('  These orders should NOT be on the HUB flat file')

    if x001_flags:
        print(f"\n{'=' * 60}")
        print(f"  🔧 X001 CORRECTIONS ({len(x001_flags)})")
        print('=' * 60)
        for f in x001_flags:
            print(f"  Order {f['order']}: {f['issue']}")

    if crate_flags:
        print(f"\n{'=' * 60}")
        print(f"  📦 CRATE STATUS MISSING ({len(crate_flags)}) — update Delivery Scheduler")
        print('=' * 60)
        for f in crate_flags:
            print(f"  {f['issue']}")

    if over_mileage_flags:
        print(f'\n  🚛 OVER {MILEAGE_MAX} MI — review mileage charge manually:')
        for o in over_mileage_flags:
            print(f'     Order {o}: {order_mileage[o]} mi')

    if zero_price_codes:
        print('\n  💲 $0.00 service codes — add rates when available:')
        for code in sorted(zero_price_codes):
            print(f'     {code}: {SERVICE_CODES[code]["description"]}')

    lookup_errors = [e for e in errors if '⚠️' in e['issue']]
    trunc_errors  = [e for e in errors if '✂️' in e['issue']]

    if lookup_errors:
        print(f"\n{'=' * 60}")
        print(f"  ⚠️  {len(lookup_errors)} ITEMS NOT FOUND — review before uploading")
        print('=' * 60)
        for e in lookup_errors:
            print(f"  Order {e['order']} | {e['model']} → {e['issue']}")

    if trunc_errors:
        print(f"\n  ✂️  {len(trunc_errors)} descriptions truncated:")
        for e in trunc_errors:
            print(f"  Order {e['order']} | {e['model']}: {e['issue']}")

    if not errors and not x001_flags:
        print('\n  ✅ Zero errors — clean run!')

    # ══════════════════════════════════════════════════════════════
    # STEP 10 — EXPORT FLAT FILE
    # ══════════════════════════════════════════════════════════════
    flat_filename = f"{prefix}_{delivery_date.strftime('%m.%d.%y')}.xlsx"
    if not _IS_COLAB:
        flat_filename = os.path.join(save_dir, flat_filename)

    if _IS_COLAB:
        tmpl_path = str(Path(CONFIG['template_colab_dir']) / CONFIG['template_filename'])
    else:
        tmpl_path = str(Path(__file__).parent / CONFIG['template_filename'])

    try:
        wb_tmpl = load_workbook(tmpl_path)
        wb      = Workbook()
        wb.remove(wb.active)
        wb.create_sheet('OrderData')
        ws_data = wb.create_sheet('data')
        for data_row in wb_tmpl['data'].iter_rows(values_only=True):
            ws_data.append(list(data_row))
        wb_tmpl.close()
        print('  ✅ Template loaded')
    except Exception as exc:
        print(f'  ⚠️  Template not found, using blank workbook: {exc}')
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet('OrderData')
        wb.create_sheet('data')

    ws        = wb['OrderData']
    flat_font = Font(name='Aptos Narrow', size=11)

    for col_idx, col_name in enumerate(FLAT_COLS, 1):
        ws.cell(row=1, column=col_idx, value=col_name).font = Font(name='Aptos Narrow', bold=False)

    for row_idx, (_, data_row) in enumerate(df_output.iterrows(), 2):
        for col_idx, col_name in enumerate(FLAT_COLS, 1):
            if col_name.startswith('_'):
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            val  = data_row.get(col_name, '')
            cell.value = None if (val == '' or (isinstance(val, float) and pd.isna(val))) else val
            cell.font  = flat_font
            if col_name == 'delivery_date':
                cell.number_format = 'mm-dd-yy'
            if col_name == 'delivery_address_1':
                cell.number_format = '@'

    wb.save(flat_filename)

    # ══════════════════════════════════════════════════════════════
    # FINANCIAL REPORT
    # ══════════════════════════════════════════════════════════════
    if has_orders_detail:
        print('\n' + '=' * 60)
        print('  FINANCIAL REPORT — REAL PRICING')
        print('=' * 60)

        fin_rows, svc_rows, product_rows = [], [], []
        _non_pieces = set(PARTS_LOOKUP.keys()) | {'WATERLINE', 'MEMO', 'TPI', 'REDEL'}

        for order_num_fmt in df_output['order_number'].unique():
            base_order   = order_num_fmt.split('-')[0]
            order_detail = df_orders[df_orders['Order #'] == base_order]
            if order_detail.empty:
                continue

            stop_name = df_output[df_output['order_number'] == order_num_fmt]['delivery_name'].iloc[0]
            city      = df_output[df_output['order_number'] == order_num_fmt]['delivery_city'].iloc[0]

            fuel_cost      = order_fuel_charge.get(base_order, 0.0)
            mileage_cost   = order_mileage_charge.get(base_order, 0.0)
            units          = order_units.get(base_order, 0)
            floor_cost     = order_floor_units.get(base_order, 0.0)
            threshold_cost = round(33.00 * units, 2) if units > 0 else 0.0

            svc_mask  = order_detail['Model Number'].astype(str).str.upper().str.startswith('X')
            svc_lines = order_detail[svc_mask]
            prd_lines = order_detail[~svc_mask]

            crate_status = order_crate_status.get(base_order)
            piece_count  = len(prd_lines[~prd_lines['Model Number'].astype(str).str.upper().isin(_non_pieces)]) if crate_status else 0

            if crate_status == 'IN-BOX':
                crate_code = 'X021'
                crate_desc = f'Crated Drop ({piece_count} piece{"s" if piece_count != 1 else ""} x ${CRATE_CHARGE_IN_BOX:.2f})'
                crate_cost = round(CRATE_CHARGE_IN_BOX * piece_count, 2)
            elif crate_status in ('OUT OF BOX', 'OUT OF BOX + INSTALL'):
                crate_code = 'X022'
                crate_desc = f'Uncrated Drop ({piece_count} piece{"s" if piece_count != 1 else ""} x ${CRATE_CHARGE_OUT_OF_BOX:.2f})'
                crate_cost = round(CRATE_CHARGE_OUT_OF_BOX * piece_count, 2)
            else:
                crate_code, crate_desc, crate_cost = None, None, 0.0

            internal_cost = fuel_cost + mileage_cost + threshold_cost + floor_cost + crate_cost
            svc_cost_hs   = svc_lines['Cost_clean'].sum()
            svc_sale      = svc_lines['ExtPrice_clean'].sum()
            svc_cost      = svc_cost_hs + internal_cost
            svc_profit    = svc_sale - svc_cost
            svc_pct       = (svc_profit / svc_sale * 100) if svc_sale > 0 else 0

            prd_cost   = prd_lines['Cost_clean'].sum()
            prd_sale   = prd_lines['ExtPrice_clean'].sum()
            prd_profit = prd_sale - prd_cost
            prd_pct    = (prd_profit / prd_sale * 100) if prd_sale > 0 else 0

            total_sale   = order_detail['ExtPrice_clean'].sum()
            total_cost   = order_detail['Cost_clean'].sum() + internal_cost
            total_profit = total_sale - total_cost
            profit_pct   = (total_profit / total_sale * 100) if total_sale > 0 else 0

            fin_rows.append({
                'order_number': order_num_fmt, 'customer': stop_name, 'city': city,
                'total_sale': round(total_sale, 2), 'total_cost': round(total_cost, 2),
                'profit_$': round(total_profit, 2), 'margin_%': round(profit_pct, 1),
                'svc_sale': round(svc_sale, 2), 'svc_cost': round(svc_cost, 2),
                'svc_profit_$': round(svc_profit, 2), 'svc_margin_%': round(svc_pct, 1),
                'prd_sale': round(prd_sale, 2), 'prd_cost': round(prd_cost, 2),
                'prd_profit_$': round(prd_profit, 2), 'prd_margin_%': round(prd_pct, 1),
            })

            for _, row in svc_lines.iterrows():
                ls = row['ExtPrice_clean']; lc = row['Cost_clean']
                lp = round(ls - lc, 2)
                svc_rows.append({'order_number': order_num_fmt, 'customer': stop_name,
                                  'code': row['Model Number'], 'description': row['Description'],
                                  'qty': row['Qty'], 'cost': lc, 'sale': ls,
                                  'profit_$': lp, 'margin_%': round((lp / ls * 100) if ls > 0 else 0, 1)})

            for code, desc, cost, sale in [
                ('FUEL',    'Fuel Surcharge',                                               fuel_cost,      0.0),
                ('MILEAGE', f'Mileage Charge ({order_mileage.get(base_order, 0)} mi)',     mileage_cost,   0.0),
                ('X018',    f'Threshold Charge ({units} units x $33.00)',                   threshold_cost, 0.0),
                ('FLOOR',   'Floor Charge (above 3rd floor)',                               floor_cost,     0.0),
                (crate_code, crate_desc,                                                    crate_cost,     0.0),
            ]:
                if cost and cost > 0:
                    svc_rows.append({'order_number': order_num_fmt, 'customer': stop_name,
                                     'code': code, 'description': desc,
                                     'qty': 1, 'cost': cost, 'sale': sale,
                                     'profit_$': round(-cost, 2), 'margin_%': 0.0})

            for _, row in prd_lines.iterrows():
                model_num    = str(row['Model Number']).strip()
                line_sale    = row['ExtPrice_clean']
                serial_units = get_serial_units(order_num_fmt, model_num, df_serial_alloc, delivery_date)
                if serial_units:
                    unit_sale = round(row['SalePrice_clean'], 2) if row['SalePrice_clean'] > 0 else line_sale
                    for unit_cost, inv_id in serial_units:
                        up = round(unit_sale - unit_cost, 2)
                        product_rows.append({
                            'order_number': order_num_fmt, 'customer': stop_name,
                            'model': model_num, 'description': str(row['Description'])[:50],
                            'qty': 1, 'cost': unit_cost, 'sale': unit_sale,
                            'profit_$': up, 'margin_%': round((up / unit_sale * 100) if unit_sale > 0 else 0, 1),
                            'inventory_id': inv_id, 'cost_source': 'serial',
                        })
                else:
                    std_cost = row['Cost_clean']
                    lp       = round(line_sale - std_cost, 2)
                    product_rows.append({
                        'order_number': order_num_fmt, 'customer': stop_name,
                        'model': model_num, 'description': str(row['Description'])[:50],
                        'qty': row['Qty'], 'cost': std_cost, 'sale': line_sale,
                        'profit_$': lp, 'margin_%': round((lp / line_sale * 100) if line_sale > 0 else 0, 1),
                        'inventory_id': '', 'cost_source': 'standard',
                    })

        df_fin     = pd.DataFrame(fin_rows)
        df_svc     = pd.DataFrame(svc_rows)
        df_product = pd.DataFrame(product_rows)

        if not df_fin.empty:
            print(f"\n{'Order':<22} {'Customer':<22} {'Sale':>9} {'Cost':>9} {'Profit $':>9} {'Margin%':>8} | {'Svc $':>7} {'Prd $':>7}")
            print('-' * 100)
            for _, r in df_fin.iterrows():
                print(f"{str(r['order_number']):<22} {str(r['customer'])[:21]:<22} "
                      f"${r['total_sale']:>8,.2f} ${r['total_cost']:>8,.2f} "
                      f"${r['profit_$']:>8,.2f}  {r['margin_%']:>6.1f}% | "
                      f"${r['svc_profit_$']:>6,.2f} ${r['prd_profit_$']:>6,.2f}")

            grand_sale   = df_fin['total_sale'].sum()
            grand_cost   = df_fin['total_cost'].sum()
            grand_profit = df_fin['profit_$'].sum()
            grand_pct    = (grand_profit / grand_sale * 100) if grand_sale > 0 else 0
            print(f"\n{'─' * 100}")
            print(f"  {'DAY TOTAL SALES':<42}  ${grand_sale:>8,.2f}")
            print(f"  {'DAY TOTAL COST':<42}  ${grand_cost:>8,.2f}")
            print(f"  {'DAY TOTAL PROFIT':<42}  ${grand_profit:>8,.2f}")
            print(f"  {'OVERALL MARGIN %':<42}  {grand_pct:>8.1f}%")
            print(f"  {'SERVICE PROFIT':<42}  ${df_fin['svc_profit_$'].sum():>8,.2f}")
            print(f"  {'PRODUCT PROFIT':<42}  ${df_fin['prd_profit_$'].sum():>8,.2f}")

            # Color scheme for the financial report
            BLUE   = 'BDD7EE'  # total columns
            ORANGE = 'FCE4D6'  # service columns
            GREEN  = 'E2EFDA'  # product columns
            by_stop_fills = {
                'total_sale': BLUE,  'total_cost': BLUE,  'profit_$': BLUE,  'margin_%': BLUE,
                'svc_sale': ORANGE,  'svc_cost': ORANGE,  'svc_profit_$': ORANGE, 'svc_margin_%': ORANGE,
                'prd_sale': GREEN,   'prd_cost': GREEN,   'prd_profit_$': GREEN,  'prd_margin_%': GREEN,
            }

            wb_fin = Workbook()
            wb_fin.remove(wb_fin.active)

            ws_by_stop = wb_fin.create_sheet('By Stop')
            write_financial_sheet(ws_by_stop, df_fin,
                sum_cols=['total_sale','total_cost','profit_$','svc_sale','svc_cost','svc_profit_$','prd_sale','prd_cost','prd_profit_$'],
                avg_cols=['margin_%','svc_margin_%','prd_margin_%'],
                dollar_cols=['total_sale','total_cost','profit_$','svc_sale','svc_cost','svc_profit_$','prd_sale','prd_cost','prd_profit_$'],
                header_fills=by_stop_fills)

            ws_svc = wb_fin.create_sheet('Service Detail')
            write_financial_sheet(ws_svc, df_svc,
                sum_cols=['cost','sale','profit_$'], avg_cols=['margin_%'],
                dollar_cols=['cost','sale','profit_$'], header_fills={})

            ws_prd = wb_fin.create_sheet('Product Detail')
            write_financial_sheet(ws_prd, df_product,
                sum_cols=['cost','sale','profit_$'], avg_cols=['margin_%'],
                dollar_cols=['cost','sale','profit_$'], header_fills={})

            fin_filename = f"{prefix}_Financial_{delivery_date.strftime('%m%d%Y')}.xlsx"
            if not _IS_COLAB:
                fin_filename = os.path.join(save_dir, fin_filename)
            wb_fin.save(fin_filename)

            if _IS_COLAB:
                _colab_files.download(fin_filename)
            print(f'\n  📥 Financial report: {fin_filename}')

    # ══════════════════════════════════════════════════════════════
    # DONE
    # ══════════════════════════════════════════════════════════════
    time.sleep(2)
    print(f"\n{'=' * 60}")
    print(f'  📥 Downloading: {flat_filename}')
    print(f'     Sheet 1: OrderData  ← upload to HUB')
    print('=' * 60)
    time.sleep(1)

    if _IS_COLAB:
        _colab_files.download(flat_filename)
    else:
        dest = os.path.join(content_dir, os.path.basename(flat_filename))
        shutil.copy(flat_filename, dest)
        print(f'  Saved to: {dest}')

    print('✅ DONE')


if __name__ == '__main__':
    main()
